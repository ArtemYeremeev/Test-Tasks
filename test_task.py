from openpyxl import load_workbook
from openpyxl import Workbook
import random
"""
Программа, отбирающая 100 случайных значений из 
предложенного списка и формирует новый список 
"""
# Используя openpyxl из исходного файла формата .xlsx извлекается необходимый столбец
source_data = load_workbook('test_task.xlsx')
data = source_data['test_task']
column_a = data['A']
# Создается пустой список для размещения данных
sample = []
# Используя цикл for и операцию append значения столбца переносятся в созданный список
for i in range(len(column_a)):
    cell = column_a[i].value
    sample.append(cell)
# При необходимости значения могут быть проверены
check = input('Значения получены. Хотите посмотреть? (да/нет)')
if check == 'да':
    print('Исходные данные -', sample)
else:
    pass
# Функция ввода запрашивает требуемое число значений
n = int(input('Сколько случайных значений вам требуется? -'))
# Создается новый пустой список для размещения данных
numbers = []
# С помощью random в цикле for заполняется новый список
# Если выпадает повторное ячейка, цикл повторяется до получения оригинального значения
for x in range(n):
    a = random.choice(sample)
    for x in range(len(sample)):
        if a == 'null':
            a = random.choice(sample)
        else:
            pass
    numbers.append(a)
print('Новый список успешно составлен')
variety = input('Вывести список на экран? (да/нет) -')
if variety == 'да':
    print('Ваш список -', numbers)
    # Выводится конечный результат в консоли
else:
    print('В таком случае полученные значения будут представленны\n '
          'в виде столбца в новом файле Excel')
    len_numbers = len(numbers)
    workbook = Workbook()
    worksheet = workbook.active
    for i in range(len_numbers):
        worksheet.cell(row=1+i, column=1, value=numbers[i])
    workbook.save('Result.xlsx')
    # Выводится конечный результат в новом файле Excel
print('Вывод данных завершен')
